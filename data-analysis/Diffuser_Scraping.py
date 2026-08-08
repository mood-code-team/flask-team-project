"""
IKEA 디퓨저(및 기타) 상품 수집기 - (색상/옵션 자동 분리 수집 기능 포함)

[기능 요약]
- 카테고리에서 기본 100개 상품 수집
- 상품 상세페이지 진입 시, 해당 상품의 '다른 색상/스타일' 버튼 링크를 찾아냄
- 새로운 색상 URL이 발견되면 수집 대기열(Queue)에 자동으로 추가하여 별도의 상품으로 수집
- 다운로드한 이미지를 엑셀 파일에 직접 삽입

[설치 명령어]
pip install pandas openpyxl pillow requests playwright
playwright install chromium
"""

import json             # 웹페이지 숨겨진 데이터(JSON-LD)를 읽기 위한 모듈
import os               # 폴더 생성, 파일 경로 설정 등을 위한 운영체제 모듈
import re               # 문자열에서 특정 패턴(특수문자 등)을 찾고 바꾸는 정규표현식 모듈
import time             # 웹사이트 과부하를 막기 위해 잠시 기다리는(sleep) 기능을 제공
from io import BytesIO  # 다운로드한 이미지 데이터를 메모리에 임시로 올려두기 위한 모듈
from typing import Any, Tuple, List, Dict # 코드 작성 시 변수의 '타입(종류)'을 명확히 지정해주는 모듈
from urllib.parse import urljoin, urlparse # 불완전한 인터넷 주소(URL)를 완전하게 조립해주는 모듈

import pandas as pd         # 엑셀, CSV 같은 표 형태의 데이터를 쉽게 다루게 해주는 강력한 라이브러리
import requests             # 파이썬에서 인터넷 주소로 접속해 이미지 파일 등을 다운로드하는 도구
from openpyxl import Workbook # 파이썬으로 엑셀 파일(.xlsx)을 새로 만들고 다루는 라이브러리
from openpyxl.drawing.image import Image as ExcelImage # 엑셀 셀 안에 그림(이미지)을 집어넣기 위한 모듈
from openpyxl.styles import Alignment, Font, PatternFill # 엑셀의 글꼴, 배경색, 가운데 정렬 등을 꾸미는 기능
from openpyxl.utils import get_column_letter # 엑셀의 열 번호(1, 2, 3...)를 알파벳(A, B, C...)으로 바꿔주는 기능
from PIL import Image as PillowImage         # 파이썬에서 이미지의 크기를 줄이거나 포맷을 바꾸는(JPG 변환) 라이브러리
from playwright.sync_api import sync_playwright # 사람이 크롬 브라우저를 직접 클릭하듯 자동으로 조종해주는 크롤링 툴


# ==================================================
# 1. 기본 설정 (상수 정의)
# ==================================================

CATEGORY_NAME = "DIFFUSER" # 데이터베이스에 들어갈 카테고리 분류명

# 이케아 홈프래그런스(디퓨저 포함) 카테고리 웹 주소
CATEGORY_URL = (
    "https://www.ikea.com/kr/ko/cat/"
    "home-fragrance-42926/"
)

MAX_PRODUCTS = 100                                    # 목표 수집 개수
OUTPUT_FOLDER = "output_diffuser"                # 결과물들이 저장될 메인 폴더
IMAGE_FOLDER = os.path.join(OUTPUT_FOLDER, "images")  # 다운받은 상품 사진들이 저장될 하위 폴더
OUTPUT_EXCEL_FILE = "diffuser_100_products.xlsx" # 최종 엑셀 파일명
OUTPUT_CSV_FILE = "diffuser_100_products.csv"    # 최종 CSV 파일명
CHECKPOINT_FILE = "diffuser_checkpoint.csv"      # 중간중간 에러를 대비해 임시로 저장해두는 파일명

# 엑셀 안에 들어갈 이미지의 가로/세로 픽셀 크기
IMAGE_WIDTH = 120
IMAGE_HEIGHT = 120

# 웹사이트가 우리를 로봇(매크로)으로 오해하고 차단하지 않도록, 평범한 크롬 브라우저인 척 속이는 신분증(헤더)
REQUEST_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/150.0.0.0 Safari/537.36"
    ),
    "Referer": "https://www.ikea.com/",
}


# ==================================================
# 2. 메타데이터(JSON-LD) 파싱 및 텍스트 정제
# ==================================================

def find_product_data(data: Any) -> dict | None:
    """웹페이지 소스코드에 숨겨져 있는 복잡한 JSON 데이터 중에서 '@type'이 'Product'(상품)인 핵심 알맹이만 찾아냅니다."""
    if isinstance(data, dict):
        data_type = data.get("@type")
        if data_type == "Product":
            return data
        if isinstance(data_type, list) and "Product" in data_type:
            return data
        # 하위 내용들을 파고들며(재귀 탐색) 계속 찾음
        for value in data.values():
            result = find_product_data(value)
            if result:
                return result
    elif isinstance(data, list):
        for item in data:
            result = find_product_data(item)
            if result:
                return result
    return None

def clean_price(value: Any) -> int:
    """'₩ 15,000' 같이 글자와 섞여 있는 가격에서 숫자만 쏙 뽑아내어 정수(int)로 만들어줍니다."""
    if value is None:
        return 0
    # 문자열을 한 글자씩 검사해서 숫자(isdigit)인 경우만 이어 붙임
    numbers_only = "".join(char for char in str(value) if char.isdigit())
    return int(numbers_only) if numbers_only else 0


# ==================================================
# 3. 이미지 URL 정제 및 추출 로직
# ==================================================

def normalize_url(value: Any, base_url: str) -> str:
    """'/images/a.jpg' 같은 불완전한 경로를 'https://www.ikea.com/images/a.jpg' 형태로 완벽하게 만들어줍니다."""
    if not isinstance(value, str): return ""
    value = value.strip()
    if not value or value.startswith("data:"): return ""
    if value.startswith("//"): return f"https:{value}"
    return urljoin(base_url, value)

def looks_like_image_url(url: str) -> bool:
    """해당 주소가 상품 상세페이지 주소가 아니라 정말 '이미지 파일'을 가리키는 주소인지 검증합니다."""
    if not url: return False
    parsed = urlparse(url)
    path = parsed.path.lower()
    # 경로에 /p/가 있고 /images/가 없으면 상품 페이지 링크이므로 False
    if "/p/" in path and "/images/" not in path: return False
    
    # 대표적인 이미지 파일 확장자들
    image_extensions = (".jpg", ".jpeg", ".png", ".webp", ".avif", ".gif")
    # 경로에 이미지 관련 단어가 있거나 확장자로 끝나면 정상적인 이미지로 판단
    return "/images/" in path or "/image/" in path or path.endswith(image_extensions)

def extract_image_from_value(image_value: Any, base_url: str) -> str:
    """숨겨진 데이터(JSON) 안에서 이미지가 문자열, 리스트, 딕셔너리 중 어떤 형태로 들어있든 악착같이 주소를 뽑아냅니다."""
    if isinstance(image_value, str):
        candidate = normalize_url(image_value, base_url)
        return candidate if looks_like_image_url(candidate) else ""
    if isinstance(image_value, list):
        for item in image_value:
            candidate = extract_image_from_value(item, base_url)
            if candidate: return candidate
        return ""
    if isinstance(image_value, dict):
        image_keys = ("contentUrl", "url", "thumbnailUrl", "src", "image")
        for key in image_keys:
            if key in image_value:
                candidate = extract_image_from_value(image_value.get(key), base_url)
                if candidate: return candidate
    return ""

def get_meta_content(page, selector: str) -> str:
    """SNS(카톡, 페이스북 등) 공유용으로 설정된 예쁜 썸네일 이미지를 찾기 위해 HTML 메타 태그를 뒤집니다."""
    locator = page.locator(selector)
    return locator.first.get_attribute("content") or "" if locator.count() > 0 else ""

def get_srcset_largest_url(srcset: str, base_url: str) -> str:
    """반응형 웹에서 화면 크기별로 여러 장 제공되는 이미지(srcset) 중 가장 고화질인 마지막 이미지를 선택합니다."""
    if not srcset: return ""
    items = [item.strip() for item in srcset.split(",") if item.strip()]
    for item in reversed(items): # 뒤에서부터 찾음 (보통 뒤쪽이 고해상도)
        candidate = normalize_url(item.split()[0], base_url)
        if looks_like_image_url(candidate): return candidate
    return ""

def collect_product_image_url(page, product_data: dict, product_url: str) -> str:
    """모든 수단을 동원해 상품의 대표 이미지 고화질 링크 1개를 무조건 찾아내는 함수입니다."""
    
    # 1단계: 상품 숨겨진 데이터(JSON)에서 찾기
    image_url = extract_image_from_value(product_data.get("image"), product_url)
    if image_url: return image_url

    # 2단계: SNS 공유용 정보(오픈그래프, 트위터 카드 등)에서 찾기
    meta_selectors = (
        "meta[property='og:image']", "meta[property='og:image:secure_url']",
        "meta[name='twitter:image']", "meta[name='twitter:image:src']",
    )
    for selector in meta_selectors:
        candidate = normalize_url(get_meta_content(page, selector), product_url)
        if looks_like_image_url(candidate): return candidate

    # 3단계: 최후의 수단으로 현재 화면에 떠있는 상품 이미지(img 태그)들을 싹 다 조사해서 찾기
    image_locators = page.locator(
        "main img[src*='/images/'], main img[srcset*='/images/'], "
        "img[src*='/images/products/'], img[srcset*='/images/products/']"
    )
    for index in range(min(image_locators.count(), 30)):
        image = image_locators.nth(index)
        candidate = normalize_url(image.get_attribute("src") or "", product_url)
        if looks_like_image_url(candidate): return candidate
        candidate = get_srcset_largest_url(image.get_attribute("srcset") or "", product_url)
        if candidate: return candidate

    return ""


# ==================================================
# 4. 카테고리 목록 페이지 자동 조종 및 링크 수집
# ==================================================

def accept_cookie_if_visible(page) -> None:
    """이케아 첫 접속 시 화면을 가리는 '쿠키 동의' 팝업창을 찾아서 자동으로 클릭해 닫아버립니다."""
    for name in ("모두 허용", "모두 동의", "Accept all", "Accept all cookies"):
        try:
            button = page.get_by_role("button", name=re.compile(name, re.IGNORECASE))
            if button.count() > 0 and button.first.is_visible():
                button.first.click(timeout=2500)
                page.wait_for_timeout(800)
                return
        except Exception: continue

def click_show_more_if_visible(page) -> bool:
    """스크롤을 내리다 보면 나오는 '더 보기' 버튼을 찾아 클릭하여 숨겨진 다음 상품들을 계속 불러옵니다."""
    for pattern in (r"더\s*보기", r"제품\s*더\s*보기", r"결과\s*더\s*보기", r"Show\s*more", r"Load\s*more"):
        try:
            buttons = page.get_by_role("button", name=re.compile(pattern, re.IGNORECASE))
            for index in range(min(buttons.count(), 5)):
                button = buttons.nth(index)
                if button.is_visible():
                    button.scroll_into_view_if_needed() # 버튼이 있는 곳까지 마우스 스크롤을 내림
                    button.click(timeout=4000)          # 버튼 클릭
                    page.wait_for_timeout(1600)         # 로딩될 때까지 잠시 대기
                    return True
        except Exception: continue
    return False

def collect_product_links(page) -> List[str]:
    """카테고리 메인 화면을 위아래로 훑으면서 목표 개수(100개)만큼 상품들의 상세페이지 주소를 쭉 수집해옵니다."""
    print("IKEA 카테고리 페이지를 여는 중입니다:", CATEGORY_URL)
    page.goto(CATEGORY_URL, wait_until="domcontentloaded", timeout=90000)
    page.wait_for_timeout(5000)
    accept_cookie_if_visible(page) # 방해되는 쿠키창 끄기

    product_urls: List[str] = []
    previous_count, unchanged_count = 0, 0

    # 무한 루프를 막기 위해 최대 80번까지만 더보기/스크롤을 시도합니다.
    for load_index in range(80):
        links = page.locator("a[href*='/p/']") # 상품 링크는 보통 /p/ 가 들어감
        
        for index in range(links.count()):
            href = links.nth(index).get_attribute("href")
            if not href: continue
            # 쓰레기 값(?이후 파라미터 등)을 잘라내고 순수한 상품 주소만 만듦
            href = urljoin(CATEGORY_URL, href).split("?")[0].split("#")[0]
            # 아직 안 가져온 새 링크면 추가
            if href not in product_urls:
                product_urls.append(href)
            if len(product_urls) >= MAX_PRODUCTS: break

        current_count = len(product_urls)
        print(f"목록 불러오기 {load_index + 1}회: 상품 링크 {current_count}개 수집됨")
        if current_count >= MAX_PRODUCTS: break # 100개 다 찾았으면 종료

        # 새로 불러왔는데 개수가 안 늘었으면 실패 카운트 증가
        unchanged_count = unchanged_count + 1 if current_count == previous_count else 0
        previous_count = current_count

        # '더 보기' 버튼이 없으면 강제로 휠을 굴려서 로딩 유도
        if not click_show_more_if_visible(page):
            page.mouse.wheel(0, 2800)
            page.wait_for_timeout(1400)

        # 7번이나 시도했는데도 개수가 안 늘면 상품이 다 떨어진 것으로 판단하고 종료
        if unchanged_count >= 7:
            print("더 이상 새로운 상품이 로드되지 않습니다.")
            break

    return product_urls[:MAX_PRODUCTS] # 목표 개수까지만 딱 잘라서 반환


# ==================================================
# 5. 상품 상세정보 및 ★색상/변형 옵션 자동 탐지★
# ==================================================

def collect_product_detail(page, product_url: str) -> Tuple[Dict | None, List[str]]:
    """
    상품 1개의 상세페이지에 접속해서 가격/설명 등을 긁어오는 동시에,
    화면에 다른 색상(옵션) 버튼이 있는지 스캔해서 그 링크들도 같이 반환합니다.
    """
    try:
        page.goto(product_url, wait_until="domcontentloaded", timeout=60000)
        page.wait_for_timeout(2500)

        # --- [핵심 기능] 화면에 보이는 '다른 색상' 버튼들의 주소 훔쳐오기 ---
        variant_urls = []
        try:
            # 이케아의 색상 선택 버튼들에 주로 쓰이는 HTML 클래스나 속성들을 모두 찾음
            style_links = page.locator(
                "div.pip-product-styles__items a, "
                "div[data-testid='pip-styles'] a, "
                "a[data-testid='pip-styles-item']"
            )
            # 찾아낸 버튼들의 링크를 하나씩 추출해서 저장
            for i in range(style_links.count()):
                href = style_links.nth(i).get_attribute("href")
                if href and "/p/" in href:
                    v_url = urljoin(product_url, href).split("?")[0].split("#")[0]
                    # 중복되지 않은 새 옵션 링크면 리스트에 추가
                    if v_url not in variant_urls:
                        variant_urls.append(v_url)
        except Exception:
            pass # 색상 버튼이 아예 없거나 에러가 나도 메인 상품 수집은 진행해야 하므로 패스
        # --------------------------------------------------------------------

        # 이제 메인 상품의 정보(이름, 가격 등) 추출 시작
        scripts = page.locator("script[type='application/ld+json']")
        for index in range(scripts.count()):
            script_text = scripts.nth(index).text_content()
            if not script_text: continue

            try: json_data = json.loads(script_text)
            except json.JSONDecodeError: continue

            product_data = find_product_data(json_data)
            if not product_data: continue

            offer = product_data.get("offers", {})
            if isinstance(offer, list): offer = offer[0] if offer else {}
            if not isinstance(offer, dict): offer = {}

            product_name = product_data.get("name", "")
            description = product_data.get("description", "")
            image_url = collect_product_image_url(page, product_data, product_url)
            price = clean_price(offer.get("price"))
            currency = offer.get("priceCurrency", "KRW")
            availability = offer.get("availability", "")
            sku = product_data.get("sku") or product_data.get("productID") or ""
            
            brand_value = product_data.get("brand", "IKEA")
            brand = brand_value.get("name", "IKEA") if isinstance(brand_value, dict) else str(brand_value)

            # 수집된 모든 정보를 하나의 딕셔너리(사전) 묶음으로 만듦
            product_info = {
                "external_source": "IKEA",
                "external_id": sku,
                "category_code": CATEGORY_NAME,
                "product_name": product_name,
                "brand": brand,
                "description": description,
                "price": price,
                "currency": currency,
                "stock_status": availability,
                "image_preview": "",          # 엑셀 사진 들어갈 빈 자리
                "thumbnail_url": image_url,
                "local_image_path": "",       # 다운로드 후 사진 파일 경로가 적힐 자리
                "source_url": product_url,
                "use_product": "검토",
                "mood_code": ""
            }
            # 알아낸 상품 정보와, 위에서 찾았던 다른 색상 링크 리스트를 동시에 반환
            return product_info, variant_urls

        return None, variant_urls

    except Exception as error:
        print("상품 상세페이지 로드 실패:", error)
        return None, []


# ==================================================
# 6. 다운로드 및 엑셀/CSV 파일 저장 시스템
# ==================================================

def safe_filename(value: str) -> str:
    """파일 이름으로 쓸 수 없는 특수문자(\, ?, : 등)를 언더바(_)로 안전하게 교체합니다."""
    value = re.sub(r'[\\/:*?"<>|]', "_", value)
    return re.sub(r"\s+", "_", value).strip("_")[:100] or "product"

def download_product_image(product: dict) -> str:
    """찾아낸 썸네일 인터넷 주소를 타고 들어가서 실제 사진 파일을 내 컴퓨터(로컬) 폴더에 저장합니다."""
    image_url = product.get("thumbnail_url", "")
    if not image_url: return ""

    os.makedirs(IMAGE_FOLDER, exist_ok=True)
    # 파일명은 유일한 식별자인 상품코드(external_id)로 설정
    base_name = safe_filename(str(product.get("external_id") or product.get("product_name") or "product"))
    image_path = os.path.join(IMAGE_FOLDER, f"{base_name}.jpg")

    # 이미 같은 이름의 사진이 멀쩡히 저장되어 있으면 중복 다운로드하지 않고 건너뜀
    if os.path.exists(image_path) and os.path.getsize(image_path) > 0:
        return image_path

    try:
        # 사진 다운로드
        response = requests.get(image_url, headers=REQUEST_HEADERS, timeout=30)
        response.raise_for_status()
        # 다운받은 이미지를 열어서 크기를 120x120(3배수)로 줄이고 JPEG로 압축하여 예쁘게 저장
        image = PillowImage.open(BytesIO(response.content)).convert("RGB")
        image.thumbnail((IMAGE_WIDTH * 3, IMAGE_HEIGHT * 3))
        image.save(image_path, format="JPEG", quality=88, optimize=True)
        return image_path # 컴퓨터에 저장된 사진 파일의 경로를 반환
    except Exception:
        return ""

# 데이터베이스 열(Column) 순서 고정
COLUMN_ORDER = [
    "external_source", "external_id", "category_code", "product_name",
    "brand", "description", "price", "currency", "stock_status",
    "image_preview", "thumbnail_url", "local_image_path",
    "source_url", "use_product", "mood_code",
]

def make_dataframe(products: List[Dict]) -> pd.DataFrame:
    """흩어져 있는 딕셔너리 데이터를 판다스(Pandas) 표 형태로 깔끔하게 조립하고 중복을 없앱니다."""
    df = pd.DataFrame(products) if products else pd.DataFrame(columns=COLUMN_ORDER)
    for col in COLUMN_ORDER:
        if col not in df.columns: df[col] = "" # 없는 열은 빈칸으로 채움
    # 상품 코드와 URL이 똑같은 쌍둥이 데이터가 있으면 쳐냄
    df = df.drop_duplicates(subset=["external_id", "source_url"], keep="first")
    return df.reindex(columns=COLUMN_ORDER) # 지정한 순서대로 열 배치

def save_checkpoint(products: List[Dict]) -> None:
    """만약 컴퓨터가 꺼지더라도 데이터가 날아가지 않도록 10개마다 몰래 CSV로 백업해두는 함수입니다."""
    if not products: return
    os.makedirs(OUTPUT_FOLDER, exist_ok=True)
    df = make_dataframe(products)
    df.to_csv(os.path.join(OUTPUT_FOLDER, CHECKPOINT_FILE), index=False, encoding="utf-8-sig")

def save_excel_with_images(dataframe: pd.DataFrame, excel_path: str) -> None:
    """수집된 글자 데이터들과 함께, 로컬에 저장된 사진들을 엑셀 셀(image_preview) 안에 쏙쏙 박아넣어 최종 엑셀을 만듭니다."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "products"

    headers = list(dataframe.columns)
    worksheet.append(headers) # 첫 번째 줄(헤더) 작성
    
    # 헤더 행 디자인 (연녹색 배경, 굵은 글씨)
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAD3")
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    image_column = headers.index("image_preview") + 1

    # 데이터를 한 줄(Row)씩 읽어서 엑셀에 채워 넣음
    for row_number, row in enumerate(dataframe.itertuples(index=False), start=2):
        values = list(row)
        values[image_column - 1] = "" # 사진이 들어갈 자리는 글자를 비워둠
        for col_num, value in enumerate(values, start=1):
            worksheet.cell(row=row_number, column=col_num, value=value)

        # 사진 파일이 내 컴퓨터에 잘 저장되어 있는지 확인
        local_image_path = str(getattr(row, "local_image_path", "") or "")
        if local_image_path and os.path.exists(local_image_path):
            try:
                # 엑셀용 이미지 객체로 만들고, 크기를 지정한 뒤 셀 안에 삽입
                excel_image = ExcelImage(local_image_path)
                excel_image.width, excel_image.height = IMAGE_WIDTH, IMAGE_HEIGHT
                worksheet.add_image(excel_image, f"{get_column_letter(image_column)}{row_number}")
                worksheet.row_dimensions[row_number].height = 95 # 사진이 찌그러지지 않게 셀 높이를 넉넉히 벌림
            except Exception: pass

    worksheet.freeze_panes = "A2" # 1행 헤더 고정 (스크롤해도 계속 보이게)
    worksheet.sheet_view.showGridLines = False # 깔끔해 보이게 엑셀 회색 격자 숨김

    # 보기 좋게 열 넓이 자동 조정
    widths = { "product_name": 48, "description": 65, "stock_status": 24, "image_preview": 20, "thumbnail_url": 65, "local_image_path": 45, "source_url": 65 }
    for i, h in enumerate(headers, 1):
        worksheet.column_dimensions[get_column_letter(i)].width = widths.get(h, 16)
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        for cell in row: cell.alignment = Alignment(vertical="center", wrap_text=True)

    workbook.save(excel_path) # 다 꾸며진 엑셀 파일 저장

def save_files(products: List[Dict]) -> None:
    """마지막으로 CSV 파일과 사진이 든 엑셀 파일을 모두 예쁘게 저장하고 요약 메시지를 띄웁니다."""
    if not products: return
    os.makedirs(OUTPUT_FOLDER, exist_ok=True); os.makedirs(IMAGE_FOLDER, exist_ok=True)
    excel_path = os.path.join(OUTPUT_FOLDER, OUTPUT_EXCEL_FILE)
    csv_path = os.path.join(OUTPUT_FOLDER, OUTPUT_CSV_FILE)
    
    df = make_dataframe(products)
    df.to_csv(csv_path, index=False, encoding="utf-8-sig")
    save_excel_with_images(df, excel_path)
    
    print("\n" + "-"*60)
    print(f"✅ 데이터 저장 완료 (총 {len(df)}개)")
    print(f"엑셀: {excel_path}\nCSV: {csv_path}")
    print("-"*60)


# ==================================================
# 7. 메인 실행부 ★(큐 대기열 방식을 활용한 색상 분리 수집 구현)★
# ==================================================

def main() -> None:
    collected_products: List[Dict] = []
    visited_urls = set() # 한 번 들어갔던 주소는 또 들어가지 않도록 기록장(Set)을 만듦

    # 자동화 브라우저(Playwright) 실행
    with sync_playwright() as playwright:
        # 화면을 띄워서(headless=False) 크롬 브라우저를 켬
        browser = playwright.chromium.launch(channel="chrome", headless=False)
        page = browser.new_page(viewport={"width": 1440, "height": 1000})

        try:
            # 1. 카테고리 메인에서 기본 상품 주소들을 긁어와서 대기열(queue)에 넣습니다.
            initial_urls = collect_product_links(page)
            queue = initial_urls.copy() 
            
            print(f"\n기본 상품 링크 {len(queue)}개를 찾았습니다.")
            print("색상 옵션이 발견되면 큐(Queue)에 자동으로 추가됩니다.\n")

            count = 1
            # 2. 대기열(queue)에 수집할 곳이 남아있고, 아직 100개를 못 채웠다면 무한 반복
            while queue and len(collected_products) < MAX_PRODUCTS:
                
                # 대기열 맨 앞에서 주소 하나를 꺼내서(pop) 방문 준비
                product_url = queue.pop(0) 

                # 만약 예전에 이미 들어갔던(수집했던) 주소라면 쿨하게 무시하고 다음으로 넘어감
                if product_url in visited_urls:
                    continue
                
                # 새로 들어가는 곳이니 방문 기록장에 도장 찍음
                visited_urls.add(product_url)
                
                print(f"[{count}/{MAX_PRODUCTS}] 상품 수집 중")
                print(f"URL: {product_url}")

                # 3. 상세페이지 진입! (상품 정보와 '다른 색상 링크들'을 동시에 들고 나옴)
                product, new_color_variants = collect_product_detail(page, product_url)

                # 상품 정보 수집에 성공했다면
                if product:
                    # 사진부터 다운로드
                    local_image_path = download_product_image(product)
                    product["local_image_path"] = local_image_path
                    
                    # 수집 완료 명단에 추가
                    collected_products.append(product)
                    print(f"  → 완료: {product.get('product_name', '')}")
                    count += 1

                    # ★ 4. 상세페이지 안에서 발견된 '다른 색상 옵션' 처리 ★
                    for variant_url in new_color_variants:
                        # 내가 아직 안 가본 주소이고, 현재 대기열에도 없는 완전 새로운 놈이라면?
                        if variant_url not in visited_urls and variant_url not in queue:
                            queue.append(variant_url) # 대기열 맨 뒤에 쏙 끼워 넣음! (나중에 방문하게 됨)
                            print(f"  [!] 새로운 색상 발견 (대기열 추가): {variant_url}")

                    # 10개마다 안전하게 임시 백업 저장
                    if len(collected_products) % 10 == 0:
                        save_checkpoint(collected_products)
                        print(f"  >>> 중간 저장 완료: {len(collected_products)}개")
                else:
                    print("  → 수집 실패 또는 데이터 누락")

                time.sleep(1.2) # 이케아 서버가 화내지 않도록 페이지 넘기기 전 잠시 휴식

        finally:
            # 작업이 끝났거나 중간에 에러가 터져도, 모아둔 건 임시 저장하고 브라우저를 안전하게 끔
            save_checkpoint(collected_products)
            browser.close()

    # 모든 작업이 끝나면 아름다운 엑셀/CSV로 출력
    save_files(collected_products)

# 파이썬에서 이 파일을 직접 실행시켰을 때만 작동하라는 표준 명령어
if __name__ == "__main__":
    main()