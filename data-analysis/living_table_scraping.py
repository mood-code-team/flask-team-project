"""
IKEA 거실/사이드 테이블 상품 수집기 - (검색 결과 기반 & 색상 옵션 수집 지원)

[카테고리 구성]
- '거실테이블', '사이드테이블', '커피테이블', '콘솔' 검색 결과를 순회하며 목표 개수를 채웁니다.
- 소파 앞 대형 테이블부터 침대/소파 옆 소형 테이블까지 완벽하게 수집합니다.
"""

import json
import os
import re
import time
from io import BytesIO
from typing import Any, Tuple, List, Dict
from urllib.parse import urljoin, urlparse, quote

import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image as PillowImage
from playwright.sync_api import sync_playwright

# ==================================================
# 1. 기본 설정 (★거실/사이드 테이블 맞춤형 검색 URL 세팅★)
# ==================================================

CATEGORY_NAME = "LIVING_ROOM_TABLE"

# 다양한 거실용 테이블을 놓치지 않기 위해 4가지 검색어를 순회합니다.
CATEGORY_URLS = [
    f"https://www.ikea.com/kr/ko/search/?q={quote('사이드테이블')}",
    f"https://www.ikea.com/kr/ko/search/?q={quote('거실테이블')}",
    f"https://www.ikea.com/kr/ko/search/?q={quote('커피테이블')}",
    f"https://www.ikea.com/kr/ko/search/?q={quote('콘솔')}"
]

MAX_PRODUCTS = 100
OUTPUT_FOLDER = "output_ikea_living_table"
IMAGE_FOLDER = os.path.join(OUTPUT_FOLDER, "images")
OUTPUT_EXCEL_FILE = "ikea_living_table_100_products.xlsx"
OUTPUT_CSV_FILE = "ikea_living_table_100_products.csv"
CHECKPOINT_FILE = "ikea_living_table_checkpoint.csv"

IMAGE_WIDTH = 120
IMAGE_HEIGHT = 120

REQUEST_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/150.0.0.0 Safari/537.36"
    ),
    "Referer": "https://www.ikea.com/",
}

# ==================================================
# 2. 메타데이터(JSON-LD) 파싱 및 정제
# ==================================================

def find_product_data(data: Any) -> dict | None:
    if isinstance(data, dict):
        data_type = data.get("@type")
        if data_type == "Product": return data
        if isinstance(data_type, list) and "Product" in data_type: return data
        for value in data.values():
            result = find_product_data(value)
            if result: return result
    elif isinstance(data, list):
        for item in data:
            result = find_product_data(item)
            if result: return result
    return None

def clean_price(value: Any) -> int:
    if value is None: return 0
    numbers_only = "".join(char for char in str(value) if char.isdigit())
    return int(numbers_only) if numbers_only else 0

# ==================================================
# 3. 이미지 URL 정제 및 추출 로직
# ==================================================

def normalize_url(value: Any, base_url: str) -> str:
    if not isinstance(value, str): return ""
    value = value.strip()
    if not value or value.startswith("data:"): return ""
    if value.startswith("//"): return f"https:{value}"
    return urljoin(base_url, value)

def looks_like_image_url(url: str) -> bool:
    if not url: return False
    parsed = urlparse(url)
    path = parsed.path.lower()
    if "/p/" in path and "/images/" not in path: return False
    image_extensions = (".jpg", ".jpeg", ".png", ".webp", ".avif", ".gif")
    return "/images/" in path or "/image/" in path or path.endswith(image_extensions)

def extract_image_from_value(image_value: Any, base_url: str) -> str:
    if isinstance(image_value, str):
        candidate = normalize_url(image_value, base_url)
        return candidate if looks_like_image_url(candidate) else ""
    if isinstance(image_value, list):
        for item in image_value:
            candidate = extract_image_from_value(item, base_url)
            if candidate: return candidate
        return ""
    if isinstance(image_value, dict):
        image_keys = ("contentUrl", "url", "thumbnailUrl", "src", "image")
        for key in image_keys:
            if key in image_value:
                candidate = extract_image_from_value(image_value.get(key), base_url)
                if candidate: return candidate
    return ""

def get_meta_content(page, selector: str) -> str:
    locator = page.locator(selector)
    return locator.first.get_attribute("content") or "" if locator.count() > 0 else ""

def get_srcset_largest_url(srcset: str, base_url: str) -> str:
    if not srcset: return ""
    items = [item.strip() for item in srcset.split(",") if item.strip()]
    for item in reversed(items):
        candidate = normalize_url(item.split()[0], base_url)
        if looks_like_image_url(candidate): return candidate
    return ""

def collect_product_image_url(page, product_data: dict, product_url: str) -> str:
    image_url = extract_image_from_value(product_data.get("image"), product_url)
    if image_url: return image_url

    meta_selectors = (
        "meta[property='og:image']", "meta[property='og:image:secure_url']",
        "meta[name='twitter:image']", "meta[name='twitter:image:src']",
    )
    for selector in meta_selectors:
        candidate = normalize_url(get_meta_content(page, selector), product_url)
        if looks_like_image_url(candidate): return candidate

    image_locators = page.locator(
        "main img[src*='/images/'], main img[srcset*='/images/'], "
        "img[src*='/images/products/'], img[srcset*='/images/products/']"
    )
    for index in range(min(image_locators.count(), 30)):
        image = image_locators.nth(index)
        candidate = normalize_url(image.get_attribute("src") or "", product_url)
        if looks_like_image_url(candidate): return candidate
        candidate = get_srcset_largest_url(image.get_attribute("srcset") or "", product_url)
        if candidate: return candidate

    return ""

# ==================================================
# 4. 검색결과 페이지 자동 조종 및 링크 수집
# ==================================================

def accept_cookie_if_visible(page) -> None:
    for name in ("모두 허용", "모두 동의", "Accept all", "Accept all cookies"):
        try:
            button = page.get_by_role("button", name=re.compile(name, re.IGNORECASE))
            if button.count() > 0 and button.first.is_visible():
                button.first.click(timeout=2500)
                page.wait_for_timeout(800)
                return
        except Exception: continue

def click_show_more_if_visible(page) -> bool:
    for pattern in (r"더\s*보기", r"제품\s*더\s*보기", r"결과\s*더\s*보기", r"Show\s*more", r"Load\s*more"):
        try:
            buttons = page.get_by_role("button", name=re.compile(pattern, re.IGNORECASE))
            for index in range(min(buttons.count(), 5)):
                button = buttons.nth(index)
                if button.is_visible():
                    button.scroll_into_view_if_needed()
                    button.click(timeout=4000)
                    page.wait_for_timeout(1600)
                    return True
        except Exception: continue
    return False

def collect_product_links(page) -> List[str]:
    product_urls: List[str] = []

    for search_url in CATEGORY_URLS:
        if len(product_urls) >= MAX_PRODUCTS:
            break

        print(f"\n[검색결과 탐색 시작] {search_url}")
        try:
            page.goto(search_url, wait_until="domcontentloaded", timeout=90000)
            page.wait_for_timeout(5000)
            accept_cookie_if_visible(page)

            previous_count, unchanged_count = len(product_urls), 0

            for load_index in range(30):
                links = page.locator("a[href*='/p/']")
                for index in range(links.count()):
                    href = links.nth(index).get_attribute("href")
                    if not href: continue
                    href = urljoin(search_url, href).split("?")[0].split("#")[0]
                    if href not in product_urls:
                        product_urls.append(href)
                    if len(product_urls) >= MAX_PRODUCTS: break

                current_count = len(product_urls)
                print(f"목록 불러오기: 현재까지 모인 링크 {current_count}개")
                if current_count >= MAX_PRODUCTS: break

                unchanged_count = unchanged_count + 1 if current_count == previous_count else 0
                previous_count = current_count

                if not click_show_more_if_visible(page):
                    page.mouse.wheel(0, 2800)
                    page.wait_for_timeout(1400)

                if unchanged_count >= 5:
                    print("이 검색어에서는 더 이상 상품이 로드되지 않아 다음 검색어로 이동합니다.")
                    break
        except Exception as e:
            print(f"탐색 중 오류 발생: {e}")
            continue

    return product_urls[:MAX_PRODUCTS]

# ==================================================
# 5. 상품 상세정보 및 색상/옵션 자동 탐지
# ==================================================

def collect_product_detail(page, product_url: str) -> Tuple[Dict | None, List[str]]:
    try:
        page.goto(product_url, wait_until="domcontentloaded", timeout=60000)
        page.wait_for_timeout(2500)

        variant_urls = []
        try:
            style_links = page.locator(
                "div.pip-product-styles__items a, "
                "div[data-testid='pip-styles'] a, "
                "a[data-testid='pip-styles-item']"
            )
            for i in range(style_links.count()):
                href = style_links.nth(i).get_attribute("href")
                if href and "/p/" in href:
                    v_url = urljoin(product_url, href).split("?")[0].split("#")[0]
                    if v_url not in variant_urls:
                        variant_urls.append(v_url)
        except Exception:
            pass

        scripts = page.locator("script[type='application/ld+json']")
        for index in range(scripts.count()):
            script_text = scripts.nth(index).text_content()
            if not script_text: continue

            try: json_data = json.loads(script_text)
            except json.JSONDecodeError: continue

            product_data = find_product_data(json_data)
            if not product_data: continue

            offer = product_data.get("offers", {})
            if isinstance(offer, list): offer = offer[0] if offer else {}
            if not isinstance(offer, dict): offer = {}

            product_name = product_data.get("name", "")
            description = product_data.get("description", "")
            image_url = collect_product_image_url(page, product_data, product_url)
            price = clean_price(offer.get("price"))
            currency = offer.get("priceCurrency", "KRW")
            availability = offer.get("availability", "")
            sku = product_data.get("sku") or product_data.get("productID") or ""
            
            brand_value = product_data.get("brand", "IKEA")
            brand = brand_value.get("name", "IKEA") if isinstance(brand_value, dict) else str(brand_value)

            product_info = {
                "external_source": "IKEA",
                "external_id": sku,
                "category_code": CATEGORY_NAME,
                "product_name": product_name,
                "brand": brand,
                "description": description,
                "price": price,
                "currency": currency,
                "stock_status": availability,
                "image_preview": "",
                "thumbnail_url": image_url,
                "local_image_path": "",
                "source_url": product_url,
                "use_product": "검토",
                "mood_code": ""
            }
            return product_info, variant_urls

        return None, variant_urls

    except Exception as error:
        print("상품 상세페이지 로드 실패:", error)
        return None, []

# ==================================================
# 6. 다운로드 및 엑셀/CSV 파일 저장 시스템
# ==================================================

def safe_filename(value: str) -> str:
    value = re.sub(r'[\\/:*?"<>|]', "_", value)
    return re.sub(r"\s+", "_", value).strip("_")[:100] or "product"

def download_product_image(product: dict) -> str:
    image_url = product.get("thumbnail_url", "")
    if not image_url: return ""

    os.makedirs(IMAGE_FOLDER, exist_ok=True)
    base_name = safe_filename(str(product.get("external_id") or product.get("product_name") or "product"))
    image_path = os.path.join(IMAGE_FOLDER, f"{base_name}.jpg")

    if os.path.exists(image_path) and os.path.getsize(image_path) > 0:
        return image_path

    try:
        response = requests.get(image_url, headers=REQUEST_HEADERS, timeout=30)
        response.raise_for_status()
        image = PillowImage.open(BytesIO(response.content)).convert("RGB")
        image.thumbnail((IMAGE_WIDTH * 3, IMAGE_HEIGHT * 3))
        image.save(image_path, format="JPEG", quality=88, optimize=True)
        return image_path
    except Exception:
        return ""

COLUMN_ORDER = [
    "external_source", "external_id", "category_code", "product_name",
    "brand", "description", "price", "currency", "stock_status",
    "image_preview", "thumbnail_url", "local_image_path",
    "source_url", "use_product", "mood_code",
]

def make_dataframe(products: List[Dict]) -> pd.DataFrame:
    df = pd.DataFrame(products) if products else pd.DataFrame(columns=COLUMN_ORDER)
    for col in COLUMN_ORDER:
        if col not in df.columns: df[col] = ""
    df = df.drop_duplicates(subset=["external_id", "source_url"], keep="first")
    return df.reindex(columns=COLUMN_ORDER)

def save_checkpoint(products: List[Dict]) -> None:
    if not products: return
    os.makedirs(OUTPUT_FOLDER, exist_ok=True)
    df = make_dataframe(products)
    df.to_csv(os.path.join(OUTPUT_FOLDER, CHECKPOINT_FILE), index=False, encoding="utf-8-sig")

def save_excel_with_images(dataframe: pd.DataFrame, excel_path: str) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "products"

    headers = list(dataframe.columns)
    worksheet.append(headers)
    
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAD3")
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    image_column = headers.index("image_preview") + 1

    for row_number, row in enumerate(dataframe.itertuples(index=False), start=2):
        values = list(row)
        values[image_column - 1] = ""
        for col_num, value in enumerate(values, start=1):
            worksheet.cell(row=row_number, column=col_num, value=value)

        local_image_path = str(getattr(row, "local_image_path", "") or "")
        if local_image_path and os.path.exists(local_image_path):
            try:
                excel_image = ExcelImage(local_image_path)
                excel_image.width, excel_image.height = IMAGE_WIDTH, IMAGE_HEIGHT
                worksheet.add_image(excel_image, f"{get_column_letter(image_column)}{row_number}")
                worksheet.row_dimensions[row_number].height = 95
            except Exception: pass

    worksheet.freeze_panes = "A2"
    worksheet.sheet_view.showGridLines = False

    widths = { "product_name": 48, "description": 65, "stock_status": 24, "image_preview": 20, "thumbnail_url": 65, "local_image_path": 45, "source_url": 65 }
    for i, h in enumerate(headers, 1):
        worksheet.column_dimensions[get_column_letter(i)].width = widths.get(h, 16)
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        for cell in row: cell.alignment = Alignment(vertical="center", wrap_text=True)

    workbook.save(excel_path)

def save_files(products: List[Dict]) -> None:
    if not products: return
    os.makedirs(OUTPUT_FOLDER, exist_ok=True)
    os.makedirs(IMAGE_FOLDER, exist_ok=True)
    excel_path = os.path.join(OUTPUT_FOLDER, OUTPUT_EXCEL_FILE)
    csv_path = os.path.join(OUTPUT_FOLDER, OUTPUT_CSV_FILE)
    
    df = make_dataframe(products)
    df.to_csv(csv_path, index=False, encoding="utf-8-sig")
    save_excel_with_images(df, excel_path)
    
    print("\n" + "-"*60)
    print(f"✅ 데이터 저장 완료 (총 {len(df)}개)")
    print(f"엑셀: {excel_path}\nCSV: {csv_path}")
    print(f"이미지 저장 폴더: {IMAGE_FOLDER}")
    print("-"*60)

# ==================================================
# 7. 메인 실행부 (큐 대기열 방식을 활용한 검색결과/색상 수집)
# ==================================================

def main() -> None:
    collected_products: List[Dict] = []
    visited_urls = set()

    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(channel="chrome", headless=False)
        page = browser.new_page(viewport={"width": 1440, "height": 1000})

        try:
            initial_urls = collect_product_links(page)
            queue = initial_urls.copy() 
            
            print(f"\n기본 상품 링크 총 {len(queue)}개를 찾았습니다.")
            print("색상/옵션이 발견되면 큐(Queue)에 자동으로 추가됩니다.\n")

            count = 1
            while queue and len(collected_products) < MAX_PRODUCTS:
                product_url = queue.pop(0) 

                if product_url in visited_urls:
                    continue
                
                visited_urls.add(product_url)
                
                print(f"[{count}/{MAX_PRODUCTS}] 상품 수집 중")
                print(f"URL: {product_url}")

                product, new_color_variants = collect_product_detail(page, product_url)

                if product:
                    local_image_path = download_product_image(product)
                    product["local_image_path"] = local_image_path
                    
                    collected_products.append(product)
                    print(f"  → 완료: {product.get('product_name', '')}")
                    count += 1

                    for variant_url in new_color_variants:
                        if variant_url not in visited_urls and variant_url not in queue:
                            queue.append(variant_url)
                            print(f"  [!] 새로운 옵션 발견 (대기열 추가): {variant_url}")

                    if len(collected_products) % 10 == 0:
                        save_checkpoint(collected_products)
                        print(f"  >>> 중간 저장 완료: {len(collected_products)}개")
                else:
                    print("  → 수집 실패 또는 데이터 누락")

                time.sleep(1.2)

        finally:
            save_checkpoint(collected_products)
            browser.close()

    save_files(collected_products)

if __name__ == "__main__":
    main()